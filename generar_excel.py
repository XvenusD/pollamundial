from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

PERSONAS = ['Duvan', 'Camila', 'Camilo', 'Mateo', 'Miguel', 'Doc Camilo', 'Kat', 'Sandra', 'Est', 'Diana']
GRUPOS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']

# Puntajes - MODIFICA ESTOS VALORES SEGÚN TUS PUNTAJES REALES
puntajes = {
    'Duvan':       {'A': 1, 'B': 3, 'C': 6, 'D': 0, 'E': 1, 'F': 0, 'G': 3, 'H': 1, 'I': 0, 'J': 6, 'K': 1, 'L': 3},
    'Camila':      {'A': 3, 'B': 1, 'C': 1, 'D': 0, 'E': 6, 'F': 1, 'G': 0, 'H': 3, 'I': 1, 'J': 1, 'K': 0, 'L': 1},
    'Camilo':      {'A': 1, 'B': 0, 'C': 3, 'D': 1, 'E': 0, 'F': 6, 'G': 1, 'H': 0, 'I': 3, 'J': 0, 'K': 6, 'L': 1},
    'Mateo':       {'A': 0, 'B': 1, 'C': 1, 'D': 3, 'E': 0, 'F': 0, 'G': 1, 'H': 6, 'I': 0, 'J': 1, 'K': 0, 'L': 3},
    'Miguel':      {'A': 1, 'B': 0, 'C': 0, 'D': 0, 'E': 1, 'F': 3, 'G': 0, 'H': 0, 'I': 1, 'J': 0, 'K': 0, 'L': 1},
    'Doc Camilo':  {'A': 0, 'B': 1, 'C': 0, 'D': 1, 'E': 0, 'F': 0, 'G': 3, 'H': 0, 'I': 0, 'J': 0, 'K': 1, 'L': 0},
    'Kat':         {'A': 1, 'B': 0, 'C': 0, 'D': 0, 'E': 0, 'F': 1, 'G': 0, 'H': 0, 'I': 0, 'J': 0, 'K': 0, 'L': 1},
    'Sandra':      {'A': 0, 'B': 1, 'C': 0, 'D': 0, 'E': 0, 'F': 0, 'G': 0, 'H': 0, 'I': 0, 'J': 0, 'K': 0, 'L': 0},
    'Est':         {'A': 0, 'B': 1, 'C': 0, 'D': 0, 'E': 0, 'F': 0, 'G': 0, 'H': 0, 'I': 0, 'J': 0, 'K': 0, 'L': 0},
    'Diana':       {'A': 0, 'B': 1, 'C': 0, 'D': 0, 'E': 0, 'F': 0, 'G': 0, 'H': 0, 'I': 0, 'J': 0, 'K': 0, 'L': 0},
}

wb = Workbook()

# ============ HOJA 1: Puntajes ============
ws1 = wb.active
ws1.title = "Puntajes"

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="0C447C", end_color="0C447C", fill_type="solid")
header_fill_gold = PatternFill(start_color="FAC775", end_color="FAC775", fill_type="solid")
header_font_dark = Font(bold=True, size=11, color="633806")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')

# Encabezados
cell = ws1.cell(row=1, column=1, value="Persona")
cell.font = header_font
cell.fill = header_fill
cell.alignment = center_align
cell.border = thin_border

for idx, grupo in enumerate(GRUPOS, start=2):
    cell = ws1.cell(row=1, column=idx, value=grupo)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

total_col = len(GRUPOS) + 2
cell = ws1.cell(row=1, column=total_col, value="TOTAL")
cell.font = header_font_dark
cell.fill = header_fill_gold
cell.alignment = center_align
cell.border = thin_border

# Datos
for row_idx, persona in enumerate(PERSONAS, start=2):
    cell = ws1.cell(row=row_idx, column=1, value=persona)
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = thin_border

    total_puntos = 0
    for col_idx, grupo in enumerate(GRUPOS, start=2):
        puntos = puntajes.get(persona, {}).get(grupo, 0)
        total_puntos += puntos
        cell = ws1.cell(row=row_idx, column=col_idx, value=puntos)
        cell.alignment = center_align
        cell.border = thin_border
        cell.font = Font(size=11)
        
        if puntos == 0:
            cell.font = Font(size=11, color="999999")
        elif puntos == 1:
            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            cell.font = Font(size=11, color="856404")
        elif puntos == 3:
            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            cell.font = Font(size=11, color="155724")
        elif puntos == 6:
            cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            cell.font = Font(size=11, color="721C24", bold=True)

    cell = ws1.cell(row=row_idx, column=total_col, value=total_puntos)
    cell.font = Font(bold=True, size=12, color="0C447C")
    cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    cell.alignment = center_align
    cell.border = thin_border

ws1.column_dimensions['A'].width = 14
for idx, _ in enumerate(GRUPOS, start=2):
    ws1.column_dimensions[get_column_letter(idx)].width = 6
ws1.column_dimensions[get_column_letter(total_col)].width = 8

# ============ HOJA 2: Detalle ============
ws2 = wb.create_sheet("Detalle")

detalle_headers = ['Grupo', 'Equipos', 'Persona', '1° Pronóstico', '2° Pronóstico', 'Puntos']
for idx, header in enumerate(detalle_headers, start=1):
    cell = ws2.cell(row=1, column=idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

grupos_info = {
    'A': 'México, Sudáfrica, Corea del Sur, Chequia',
    'B': 'Canadá, Bosnia, Qatar, Suiza',
    'C': 'Brasil, Marruecos, Haití, Escocia',
    'D': 'EE.UU., Paraguay, Australia, Turquía',
    'E': 'Alemania, Costa de Marfil, Ecuador, Curazao',
    'F': 'Países Bajos, Japón, Suecia, Túnez',
    'G': 'Bélgica, Egipto, Irán, N. Zelanda',
    'H': 'España, Cabo Verde, Arabia Saudí, Uruguay',
    'I': 'Francia, Senegal, Irak, Noruega',
    'J': 'Argentina, Argelia, Austria, Jordania',
    'K': 'Portugal, RD Congo, Uzbekistán, Colombia',
    'L': 'Inglaterra, Croacia, Ghana, Panamá',
}

row = 2
for grupo_id, equipos in grupos_info.items():
    for persona in PERSONAS:
        ws2.cell(row=row, column=1, value=f"Grupo {grupo_id}").border = thin_border
        ws2.cell(row=row, column=2, value=equipos).border = thin_border
        ws2.cell(row=row, column=3, value=persona).border = thin_border
        ws2.cell(row=row, column=4, value="—").border = thin_border
        ws2.cell(row=row, column=5, value="—").border = thin_border
        
        puntos = puntajes.get(persona, {}).get(grupo_id, 0)
        cell = ws2.cell(row=row, column=6, value=puntos)
        cell.border = thin_border
        cell.alignment = center_align
        row += 1

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 8

# ============ HOJA 3: Ranking ============
ws3 = wb.create_sheet("Ranking")

totales = {persona: sum(puntajes.get(persona, {}).values()) for persona in PERSONAS}
ranking = sorted(totales.items(), key=lambda x: x[1], reverse=True)

ranking_headers = ['Pos', 'Persona', 'Total', 'Detalle']
for idx, header in enumerate(ranking_headers, start=1):
    cell = ws3.cell(row=1, column=idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for pos, (persona, total) in enumerate(ranking, start=1):
    ws3.cell(row=pos+1, column=1, value=pos).border = thin_border
    ws3.cell(row=pos+1, column=1).alignment = center_align
    
    cell_name = ws3.cell(row=pos+1, column=2, value=persona)
    cell_name.border = thin_border
    if pos == 1:
        cell_name.font = Font(bold=True, size=12, color="B8860B")
    elif pos == 2:
        cell_name.font = Font(bold=True, size=12, color="808080")
    elif pos == 3:
        cell_name.font = Font(bold=True, size=12, color="CD853F")
    
    cell_total = ws3.cell(row=pos+1, column=3, value=total)
    cell_total.border = thin_border
    cell_total.alignment = center_align
    cell_total.font = Font(bold=True, size=14, color="0C447C")
    
    detalle = [f"{g}:{p}" for g in GRUPOS if (p := puntajes.get(persona, {}).get(g, 0)) > 0]
    ws3.cell(row=pos+1, column=4, value=", ".join(detalle) if detalle else "—").border = thin_border

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 40

# Guardar
fecha = datetime.now().strftime("%Y%m%d")
filename = f"Polla_Mundial2026_{fecha}.xlsx"
wb.save(filename)
print(f"✅ Excel generado exitosamente: {filename}")
print(f"   📊 Hoja 1: Puntajes (formato para cargar en la web)")
print(f"   📋 Hoja 2: Detalle")
print(f"   🏆 Hoja 3: Ranking")